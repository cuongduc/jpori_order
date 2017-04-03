from __future__ import unicode_literals

import os
from openpyxl import load_workbook

from product.models import Product, ProductInstance

# Define index for each field
FIELD_IDX = {
    'product_type': 1,
    'product_group': 2,
    'id': 3,
    'name': 4,
    'selling_price': 5,
    'buying_price': 6,
    'description': 13,
    'imeis': 19
}


def import_products(excel_path=None):
    """
    Read customer from excel files and save to database
    """
    # Check if file is existed
    if not (os.path.exists(excel_path) and os.path.isfile(excel_path)):
        raise IOError("File not found")
    else:
        file = open(excel_path, "rb")
    # Load workbook for file
    wb = load_workbook(file)
    # Get the active sheet (which is often the only sheet we need)
    sheet = wb.active
    start_row = 2  # as 1 is header row
    end_row = sheet.max_row + 1
    result = []
    for i in range(start_row, end_row):
        product_type = sheet.cell(row=i, column=FIELD_IDX[
                                  'product_type']).value
        product_group = sheet.cell(row=i, column=FIELD_IDX[
                                   'product_group']).value
        id = sheet.cell(row=i, column=FIELD_IDX['id']).value
        name = sheet.cell(row=i, column=FIELD_IDX['name']).value
        selling_price = sheet.cell(row=i, column=FIELD_IDX[
                                   'selling_price']).value
        buying_price = sheet.cell(row=i, column=FIELD_IDX[
                                  'buying_price']).value
        description = sheet.cell(row=i, column=FIELD_IDX['description']).value
        imeis = sheet.cell(row=i, column=FIELD_IDX['imeis']).value

        if not check_product_exist(id=id):
            product = Product(
                id=id, name=name, description=description,
                product_type=product_type, product_group=product_group,
                selling_price=selling_price, buying_price=buying_price
            )
            product.save()
            create_product_instances(imeis, product)
            result.append(product)

    return result


def check_product_exist(id):
    try:
        product = Product.objects.get(id=id)
        return product is not None
    except Product.DoesNotExist:
        return False


def create_product_instances(imeis, product):
    # Get list of imei
    ils = process_imeis(imeis)
    print(ils)
    if len(ils) == 0:
        return
    for imei in ils:
        print(type(imei))
        if len(imei):
            ProductInstance.objects.create(product=product, imei=imei)
    return True


def process_imeis(imeis):
    if imeis is None:
        return []
    if not isinstance(imeis, int):
        return imeis.split("|")
    elif imeis != '':
        return [str(imeis)]
