from __future__ import unicode_literals

import os
from openpyxl import load_workbook


class ExcelReader:

    def __init__(self):
        self.workbook = None

    def load_excel(self, filename):
        try:
            with open(filename, 'rb') as f:
                self.workbook = load_workbook(f)
                f.close()
        except IOError as e:
            raise IOError("IO error {}: {}".format(e.errno, e.strerror))
        except:
            raise Exception("Unexpected exception when reading file!")

    def get_active_sheet(self):
        return self.workbook.active

    def get_header(self):
        ws = self.get_active_sheet()
        col_min = ws.min_column
        col_max = ws.max_column
        row_idx = 1
        header = []
        for col in range(col_min, col_max + 1):
            cell = ws.cell(row=row_idx, column=col)
            header.append(str(cell.value))
        return header

    def __str__(self):
        return "Excel reader"
