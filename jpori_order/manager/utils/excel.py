from __future__ import unicode_literals

import os
from django.conf import settings
from django.core.files.base import ContentFile

from openpyxl import load_workbook

from customer.models import Customer

# Define index for each field
FIELD_IDX = {
    'id': 1,
    'name': 2,
    'email': 3,
    'phone': 4,
    'address': 5,
    'birthday': 9,
    'note': 10
}


def import_customers(excel_path=None):
    """
    Read customer from excel files and save to database
    """
    # Check if file is existed
    if not (os.path.exists(excel_path) and os.path.isfile(excel_path)):
        raise IOError("File not found")
    else:
        file = open(excel_path, "rb")
    # Load workbook for file
    wb = load_workbook(file)
    # Get the active sheet (which is often the only sheet we need)
    sheet = wb.active
    start_row = 2  # as 1 is header row
    end_row   = sheet.max_row + 1
    result = []
    for i in range(start_row, end_row):
        id       = sheet.cell(row=i, column=FIELD_IDX['id']).value
        name     = sheet.cell(row=i, column=FIELD_IDX['name']).value
        email    = sheet.cell(row=i, column=FIELD_IDX['email']).value
        phone    = sheet.cell(row=i, column=FIELD_IDX['phone']).value
        birthday = sheet.cell(row=i, column=FIELD_IDX['birthday']).value
        address  = sheet.cell(row=i, column=FIELD_IDX['address']).value
        if not check_customer_exist(id=id):
            customer = Customer(
                id=id, name=name, email=email, phone=phone,
                birthday=birthday, address=address
            )
            customer.save()
            result.append(customer)

    return result


def check_customer_exist(id):
    """
    Check whether a customer with given id is already existed
    :param id: 
    :return: 
    """
    try:
        customer = Customer.objects.get(id=id)
        return customer is not None
    except Customer.DoesNotExist:
        return False


def save_uploaded_file(file_obj, filename):
        """
        Save uploaded file to media folder to temporary use
        :param file_obj:
        :param filename:
        :return:
        """
        path = "{}/{}".format(settings.MEDIA_ROOT, filename)
        with open(path, 'wb+') as dest:
            file_content = ContentFile(file_obj.read())
            try:
                for chunk in file_content.chunks():
                    dest.write(chunk)
                dest.close()
                return path
            except:
                raise IOError("Cannot write uploaded file.")


def delete_temp_file(filepath):
    if os.path.exists(filepath) and os.path.isfile(filepath):
        os.remove(filepath)
