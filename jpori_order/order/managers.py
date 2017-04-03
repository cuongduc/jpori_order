from __future__ import unicode_literals

import os

from openpyxl import load_workbook
from django.db import models
from .constants import ORDER_FIELD_DICT
from customer.models import Customer
from product.models import Product, ProductInstance
from product.exceptions import ProductNotFound, ProductInstanceNotFound


class OrderManager(models.Manager):
    use_for_related_fields = True

    def import_from_excel(self, excel_path=None):
        """
        Import orders from excel file
        """
        # check for valid file path
        if not (self._check_valid_file):
            raise IOError("File not found.")
        file = open(excel_path, "rb")
        # Load excel file and get active workbook
        wb = load_workbook(file)
        sheet = wb.active
        start_row = 2  # as the first row is header
        end_row = sheet.max_row
        result = []
        for i in range(start_row, end_row):
            id = self._get_cell_value(sheet, i, "id")
            time = self._get_cell_value(sheet, i, "time")
            customer_id = self._get_cell_value(sheet, i, "customer_id")
            customer_name = self._get_cell_value(sheet, i, "customer_name")
            customer_phone = self._get_cell_value(sheet, i, "customer_phone")
            salesperson = self._get_cell_value(sheet, i, "salesperson")
            creator = self._get_cell_value(sheet, i, "creator")
            description = self._get_cell_value(sheet, i, "description")
            total = self._get_cell_value(sheet, i, "total")
            discount = self._get_cell_value(sheet, i, "discount")
            product_id = self._get_cell_value(sheet, i, "product_id")
            # product_name = self._get_cell_value(sheet, i, "product_name")
            product_imei = self._get_cell_value(sheet, i, "product_imei")
            product_quantity = self._get_cell_value(
                sheet, i, "product_quantity")
            product_price = self._get_cell_value(sheet, i, "product_price")
            product_discount = self._get_cell_value(
                sheet, i, "product_discount")
            product_total = self._get_cell_value(sheet, i, "product_total")

            # Check if a product with given id and imei is existed.
            try:
                product = Product.objects.get(id=product_id)
            except Product.DoesNotExist:
                raise ProductNotFound(
                    feature=product_id,
                    message="Không tồn tại sản phẩm với id={}".format(product_id)
                )

            if (product_imei is not None and product_imei != ""):
                try:
                    product_instance = Product.objects.filter(imeis__imei=product_imei)
                except ProductInstance.DoesNotExist:
                    raise ProductInstanceNotFound(
                        feature=product_imei,
                        message="Không tồn tại IMEI={}".format(product_imei)
                    )
            # Check if customer is existed
            if (customer_id is None or customer_id == ""):
                continue
            customer = self._get_or_create_customer(id=customer_id,
                                                    name=customer_name,
                                                    phone=customer_phone)
            try:
                order = self.get(id=id)
            except self.model.DoesNotExist:
                order = self.create(id=id, time=time, salesperson=salesperson,
                                    creator=creator, description=description,
                                    total=total, discount=discount,
                                    customer_id=customer_id)
                # Create order details
                details = self.create_order_detail(order_id=order.id,
                                                   product_id=product.id,
                                                   quantity=product_quantity,
                                                   price=product_price,
                                                   discount=product_discount,
                                                   product_total=product_total,
                                                   customer_id=customer.id)
                order.details.add(details)
            result.append(order)

        return result

    def _check_valid_file(self, path):
        """
        Verify if the file is existed
        """
        return (os.path.exists(path) and os.path.isfile(path))

    def _get_cell_value(self, sheet, row_index, column_name):
        """
        Get value of a cell in given sheet
        """
        return sheet.cell(
            row=row_index,
            column=ORDER_FIELD_DICT[column_name]
        ).value

    def _check_valid_product(self, id):
        try:
            product = Product.objects.get(id=id)
            return product
        except Product.DoesNotExist:
            return None

    def _get_or_create_customer(self, id, **kwargs):
        try:
            customer = Customer.objects.get(id=id)
            return customer
        except Customer.DoesNotExist:
            customer = Customer.objects.create(id=id,
                                               name=kwargs['name'],
                                               phone=kwargs['phone'])
            return customer

    def create(self, **kwargs):
        order = self.model()
        order.id = kwargs['id']
        order.time = kwargs['time']
        order.salesperson = kwargs['salesperson']
        order.creator = kwargs['creator']
        order.description = kwargs['description']
        order.total = kwargs['total']
        order.discount = kwargs['discount']
        order.customer_id = kwargs['customer_id']
        order.save(force_insert=True, using=self.db)
        return order

    def create_order_detail(self, **kwargs):
        details = self.model.details.rel.related_model.objects.create(
            product_id=kwargs['product_id'],
            quantity=kwargs['quantity'],
            discount=kwargs['discount'],
            product_total=kwargs['product_total'],
            order_id=kwargs['order_id'],
            customer_id=kwargs['customer_id'],
        )
        print(details)
        return details
